import datetime
import os
from datetime import date

from flask import Blueprint, send_file, request
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

import db
from excel.excel_sql import sql_select_otsut
# from win32com import client
# from fpdf import FPDF
# import pandas as pd
# import numpy as np


excel = Blueprint('excel', __name__)

# date = datetime.datetime.now()
# date_for_text = date.strftime("%d.%m.%Y")


def book_create(date_start, date_finish):
    # date_start = date_start.strftime("%d.%m.%Y")
    # date_finish = date_finish.strftime("%d.%m.%Y")
    data = db.select(sql_select_otsut.format(date_start=date_start, date_finish=date_finish))
    print(sql_select_otsut.format(date_start=date_start, date_finish=date_finish))
    # sql_select = sql_select_template.format(date_start=date_start, date_finish=date_finish)
    # data = db_select(sql_select)
    book = Workbook()
    sheet = book.active
    sheet.title = "Период отсутствия персонала"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToHeight = False
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_margins.left = 0.3
    sheet.page_margins.right = 0.1
    sheet.page_margins.top = 0.4
    sheet.page_margins.bottom = 0.44
    sheet.page_margins.header = 0.2
    sheet.oddHeader.left.text = "Отчет: 'Периоды отсутствия персонала' Стр. &[Page] из &N"
    sheet.oddHeader.left.size = 12
    sheet.oddHeader.left.font = "Tahoma,Bold"
    sheet.oddHeader.left.color = "000000"
    row2 = sheet.row_dimensions[2]
    row3 = sheet.row_dimensions[3]
    row5 = sheet.row_dimensions[5]
    row6 = sheet.row_dimensions[6]
    row2.height = 30
    row3.height = 30
    row5.height = 20
    row6.height = 20
    sheet.append([" "])
    sheet.append(["Отчет по отсутствующим сотрудникам"])
    sheet.append([f"с {date_start} по {date_finish}"])
    sheet.append([" "])
    sheet.append(["ФИО", "Подразделение", "Причина", "Период"])
    sheet.append([" ", " ", " ", "Начало", "Конец"])
    sheet.merge_cells('D5:E5')
    sheet.merge_cells('A5:A6')
    sheet.merge_cells('B5:B6')
    sheet.merge_cells('C5:C6')
    # sheet.merge_cells('F5:E6')
    sheet.merge_cells('A2:E2')
    sheet.merge_cells('A3:E3')
    list_result = []
    # Построчно передача данных для разбора в функцию, после запись в СПИСОК result
    for a in data:
        data_result = convert_data(a)
        list_result.append(data_result)
    list_result.sort(key=sort_key)
    # ---------------------------------------------
    for row in list_result:
        sheet.append(row)
    # 1 стобец заливка, Шрифт жирный-----------------------------------------------------
    color_fill = PatternFill(start_color='dce6f1', end_color='dce6f1', fill_type='solid')
    for i in range(7, (sheet.max_row + 1)):
        x = "A" + str(i)
        sheet[f'{x}'].fill = color_fill
        sheet[f'{x}'].font = Font(bold=True)
    # Заливка и шрифт шапки-------------------------------------------------------------
    for i in range(1, sheet.max_column):
        sheet.cell(row=5, column=i).fill = color_fill
        sheet.cell(row=6, column=i).fill = color_fill
    # КОСТЫЛЬ----------------------------------------------------------------------
    sheet.cell(row=6, column=5).fill = color_fill
    # переформат даты в dd/mm/yyyy---------------------------------------------------
    for i in range(1, sheet.max_row):
        dateCell = sheet.cell(row=i+1, column=5)
        dateCell.number_format = 'dd/mm/yyyy;@'
    for i in range(1, sheet.max_row):
        dateCell = sheet.cell(row=i+1, column=4)
        dateCell.number_format = 'dd/mm/yyyy;@'
    # -------------------------------------------------------------------------------
    # for i in range(1, sheet.max_row):
    #     if sheet.cell(row=i+1, column=6).value == "Отсутствует":
    #         for s in range(1, sheet.max_column + 1):
    #             sheet.cell(row=i + 1, column=s).fill = PatternFill(start_color="FFC7CE", fill_type="solid")
    fullRange = "A6:" + get_column_letter(sheet.max_column) + str(sheet.max_row)
    sheet.auto_filter.ref = fullRange
    set_column_widths(sheet)
    set_border(sheet)
    set_border_heading(sheet)
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15
    # sheet.column_dimensions['F'].width = 20
    side_top = Side(border_style=None, color='FF000000')
    for i in range(1, (sheet.max_column + 1)):
        for s in range(1, 5):
            sheet.cell(row=s, column=i).border = Border(top=side_top, bottom=side_top, left=side_top, right=side_top)
    # ws.cell(row=5, column=i).font = Font(size=14, bold=True)
    # book.save(f"{date}_otchet_po_otsutstviyu.xlsx")
    book.save("excel/otchet_po_otsutstviyu.xlsx")


def convert_data(conv_data):
    list_data = list(conv_data)
    if list_data[4] is None:
        list_data[4] = ''
        # list_data.append("Отсутствует")
    else:
        list_data[4] = list_data[4].date()
        list_data[3] = list_data[3].date()

        # if list_data[4] >= date and list_data[3] <= date:
        #     list_data.append("Отсутствует")

    convert_result = tuple(list_data)
    return convert_result


# сортировка по столбцу 2-----------------------
def sort_key(people):
    return people[0]
# ---------------------------------------------


# Рамки ячеек--------------------------------------------------------------------
def set_border(ws):
    side = Side(border_style='thin', color='FF000000')

    for cell in ws._cells.values():
        cell.border = Border(top=side, bottom=side, left=side, right=side)


# Выравнивание заголовка и жирная граница---------------------------------------------------
def set_border_heading(ws):
    side_border = Side(border_style='medium', color='FF000000')
    for i in range(1, (ws.max_column + 1)):
        ws.cell(row=5, column=i).font = Font(size=14, bold=True)
        ws.cell(row=5, column=i).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=6, column=i).font = Font(size=14, bold=True)
        ws.cell(row=6, column=i).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=2, column=i).font = Font(size=14, bold=True)
        ws.cell(row=2, column=i).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=3, column=i).font = Font(size=14, bold=True)
        ws.cell(row=3, column=i).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=5, column=i).border = Border(top=side_border, bottom=side_border, left=side_border, right=side_border)
        ws.cell(row=6, column=i).border = Border(top=side_border, bottom=side_border, left=side_border, right=side_border)


# Подгон размера столбцов--------------------------------------------------------
def set_column_widths(ws):
    column_widths = []
    for row in ws.iter_rows():
        for i, cell in enumerate(row):
            try:
                column_widths[i] = max(column_widths[i], len(str(cell.value))+3)

            except IndexError:
                column_widths.append(len(str(cell.value)))

    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = column_width

# -----------------------------------------------------------------------------

@excel.route('/', methods=['GET', 'POST'])
def excel_ots():

    dtn = request.args.get('dtn')
    dtk = request.args.get('dtk')
    # date_start_example = '01.01.2021'
    # date_finish_example = '31.12.2021'
    book_create(dtn, dtk)

    return send_file("excel/otchet_po_otsutstviyu.xlsx",
                     mimetype='xlsx',
                     attachment_filename='otchet_po_otsutstviyu.xlsx',
                     as_attachment=True)



# @excel.route('/pdf', methods=['GET', 'POST'])
# def excel_to_pdf():
#     dtn = request.args.get('dtn')
#     dtk = request.args.get('dtk')
#     # dtn = '01.01.2021'
#     # dtk = '31.12.2021'
#     book_create(dtn, dtk)
#
#     # # Open Microsoft Excel
#     #
#     # excel = client.Dispatch("Excel.Application")
#     #
#     # # Read Excel File
#     # sheets = excel.Workbooks.Open("excel/otchet_po_otsutstviyu.xlsx")
#     # work_sheets = sheets.Worksheets[0]
#     #
#     # # Convert into PDF File
#     # work_sheets.ExportAsFixedFormat(0, 'excel/otchet_po_otsutstviyu.pdf')
#
#     # read in the .xlsx file just created
#     # df_2 = pd.read_excel('test.xlsx')
#
#     # my_wb = openpyxl.Workbook('excel/otchet_po_otsutstviyu.xlsx')
#     # df_2 = my_wb.active
#     #
#     #
#     #
#     #
#     # # creating a pdf in called test.pdf in the current directory
#     # pdf = FPDF()
#     # pdf.add_page()
#     # pdf.set_xy(0, 0)
#     # pdf.set_font('arial', 'B', 14)
#     # pdf.cell(60)
#     # pdf.cell(70, 10, 'Writing a PDF from python', 0, 2, 'C')
#     # pdf.cell(-40)
#     # pdf.cell(50, 10, 'Index Column', 1, 0, 'C')
#     # pdf.cell(40, 10, 'Col A', 1, 0, 'C')
#     # pdf.cell(40, 10, 'Col B', 1, 2, 'C')
#     # pdf.cell(-90)
#     # pdf.set_font('arial', '', 12)
#     # for i in range(0, len(df_2) - 1):
#     #     col_ind = str(i)
#     #     col_a = str(df_2.A.ix[i])
#     #     col_b = str(df_2.B.ix[i])
#     #     pdf.cell(50, 10, '%s' % (col_ind), 1, 0, 'C')
#     #     pdf.cell(40, 10, '%s' % (col_a), 0, 0, 'C')
#     #     pdf.cell(40, 10, '%s' % (col_b), 0, 2, 'C')
#     #     pdf.cell(-90)
#     # pdf.output('excel/otchet_po_otsutstviyu.pdf', 'F')
#     #
#     # Creating a dataframe and saving as test.xlsx in current directory
#     df_1 = pd.DataFrame(np.random.randn(10, 2), columns=list('AB'))
#     writer = pd.ExcelWriter('excel/otchet_po_otsutstviyu.xlsx')
#     df_1.to_excel(writer)
#     writer.save()
#
#     #read in the .xlsx file just created
#     df_2 = pd.read_excel('excel/otchet_po_otsutstviyu.xlsx')
#
#     #creating a pdf in called test.pdf in the current directory
#     pdf = FPDF()
#     pdf.add_page()
#     pdf.set_xy(0, 0)
#     pdf.set_font('arial', 'B', 14)
#     pdf.cell(60)
#     pdf.cell(70, 10, 'Writing a PDF from python', 0, 2, 'C')
#     pdf.cell(-40)
#     pdf.cell(50, 10, 'Index Column', 1, 0, 'C')
#     pdf.cell(40, 10, 'Col A', 1, 0, 'C')
#     pdf.cell(40, 10, 'Col B', 1, 2, 'C')
#     pdf.cell(-90)
#     pdf.set_font('arial', '', 12)
#     # for i in range(0, len(df_2)-1):
#     #     col_ind = str(i)
#     #     col_a = str(df_2.A.x[i])
#     #     col_b = str(df_2.B.ix[i])
#     #     pdf.cell(50, 10, '%s' % (col_ind), 1, 0, 'C')
#     #     pdf.cell(40, 10, '%s' % (col_a), 0, 0, 'C')
#     #     pdf.cell(40, 10, '%s' % (col_b), 0, 2, 'C')
#     #     pdf.cell(-90)
#     pdf.output('excel/otchet_po_otsutstviyu.pdf', 'F')
#
#     # pdf = FPDF()
#     #
#     # pdf.add_page()
#     #
#     # pdf.set_font("Arial", size=25)
#     #
#     # # create a cell
#     # pdf.cell(200, 10, txt="JournalDev",
#     #          ln=1, align='C')
#     #
#     # pdf.cell(200, 10, txt="Welcome to the world of technologies!",
#     #          ln=2, align='C')
#     #
#     # pdf.output("excel/otchet_po_otsutstviyu.pdf")
#
#     return send_file("excel/otchet_po_otsutstviyu.pdf",
#                      mimetype='pdf',
#                      attachment_filename='otchet_po_otsutstviyu.pdf',
#                      as_attachment=True)